from flask import Flask, render_template, request, send_file
import pandas as pd
from openpyxl import load_workbook
from docx import Document
import os
import re

app = Flask(__name__)

# Function to normalize WERS codes
def normalize_code(code):
    if code is None:
        return ""
    code = re.sub(r'[-_#]', ' ', code)  # Replace -, _, and # with spaces
    code = re.sub(r'\s*--\s*', ' ', code)  # Replace -- with a space
    code = re.sub(r'\s+', ' ', code)  # Replace multiple spaces with a single space
    return code.strip()

# Function to determine if a row corresponds to a single entry
def is_single_entry(row):
    wers_code = str(row['WERS Code'])
    # Check if the code contains only alphanumeric characters and is not empty
    return bool(wers_code and re.match(r'^[A-Za-z0-9]+$', wers_code))

# Function to extract the alphanumeric code from the end of Feature WERS Description
def extract_end_code(description):
    if not isinstance(description, str):
        return None
    # Split by any separator and get the last part
    parts = re.split(r'[-\s]+', description.strip())
    if parts:
        last_part = parts[-1]
        print(f"Description: {description} -> Last part: {last_part}")  # Debug print
        return last_part
    return None

@app.route('/')
def upload_files():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def process_files():
    uploads_dir = 'uploads'
    # Ensure the uploads directory exists
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)

    if 'excel_file' not in request.files or 'word_file' not in request.files or 'voci_excel_file' not in request.files:
        return "Missing file(s)"

    # Save uploaded files
    excel_file = request.files['excel_file']
    word_file = request.files['word_file']
    voci_excel_file = request.files['voci_excel_file']
    excel_header = int(request.form['excel_header'])
    voci_header = int(request.form['voci_header'])

    excel_path = os.path.join(uploads_dir, excel_file.filename)
    word_path = os.path.join(uploads_dir, word_file.filename)
    voci_excel_path = os.path.join(uploads_dir, voci_excel_file.filename)

    excel_file.save(excel_path)
    word_file.save(word_path)
    voci_excel_file.save(voci_excel_path)

    try:
        excel_data = pd.read_excel(excel_path, header=excel_header - 1)
    except Exception as e:
        return f"Error reading Excel file: {e}"

    # Ensure all required columns exist
    required_excel_columns = ['Feature WERS Code', 'Feature WERS Description', 'Top Family WERS Code', 'Top Family Engineering Description']
    for col in required_excel_columns:
        if col not in excel_data.columns:
            return f"Column '{col}' not found in the Excel file."

    # Create mappings
    feature_description_map = {}
    top_family_map = {}  # New mapping to store top family for each feature code
    
    print("Creating mappings...")  # Debug print
    
    # Process the Excel file to identify YZA family codes
    for idx, row in excel_data.iterrows():
        try:
            feature_code = str(row['Feature WERS Code']).strip() if pd.notna(row['Feature WERS Code']) else None
            top_family = str(row['Top Family WERS Code']).strip() if pd.notna(row['Top Family WERS Code']) else None
            description = str(row['Feature WERS Description']).strip() if pd.notna(row['Feature WERS Description']) else None
            
            if feature_code and description and top_family:
                # Check if this is a YZA family code
                if top_family == 'YZA':
                    # Store the top family mapping
                    top_family_map[feature_code] = top_family
                    
                    # Extract the end code from the description
                    end_code = extract_end_code(description)
                    if end_code:
                        feature_description_map[feature_code] = end_code
                        print(f"Mapped YZA code {feature_code} -> {end_code} from '{description}'")  # Debug print
        except Exception as e:
            print(f"Error processing row {idx}: {e}")  # Debug print
            continue
    
    print("\nFeature Description Map:", feature_description_map)  # Debug print
    print("\nTop Family Map:", top_family_map)  # Debug print

    try:
        voci_data = pd.read_excel(voci_excel_path, header=voci_header - 1)
    except Exception as e:
        return f"Error reading VOCI Excel file: {e}"

    required_columns = ['WERS Code', 'Sales Code']
    for col in required_columns:
        if col not in voci_data.columns:
            return f"Column '{col}' not found in the VOCI Excel file."

    voci_codes = voci_data[['WERS Code', 'Sales Code']].dropna().astype(str)
    
    # Create a mapping from WERS Code to Sales Code
    single_code_sales = {}
    group_code_sales = {}

    print("\nProcessing VOCI data...")  # Debug print
    # First pass: identify all single entries
    for index, row in voci_codes.iterrows():
        wers_code = normalize_code(row['WERS Code'])
        sales_code = row['Sales Code']
        
        if wers_code and is_single_entry(row):
            # Handle YZA family codes
            if wers_code in top_family_map:
                print(f"\nProcessing YZA family code: {wers_code}")
                if wers_code in feature_description_map:
                    end_code = feature_description_map[wers_code]
                    print(f"Using end code for {wers_code}: {end_code}")
                    sales_code = end_code
            single_code_sales[wers_code] = sales_code
            print(f"Added single code: {wers_code} -> {sales_code}")  # Debug print

    # Second pass: add group entries only if no single entry exists
    for index, row in voci_codes.iterrows():
        wers_code = normalize_code(row['WERS Code'])
        sales_code = row['Sales Code']
        
        if wers_code and not is_single_entry(row):
            # Handle YZA family codes
            if wers_code in top_family_map:
                print(f"\nProcessing YZA family code (group): {wers_code}")
                if wers_code in feature_description_map:
                    end_code = feature_description_map[wers_code]
                    print(f"Using end code for {wers_code}: {end_code}")
                    sales_code = end_code
            if wers_code not in single_code_sales:
                group_code_sales[wers_code] = sales_code
                print(f"Added group code: {wers_code} -> {sales_code}")  # Debug print

    # Combine all results
    results = []
    
    # Get all codes from Excel
    codes_from_excel = excel_data['Feature WERS Code'].dropna().astype(str).tolist()
    
    # Process codes found in Word document
    try:
        doc = Document(word_path)
        text_content = [paragraph.text for paragraph in doc.paragraphs]
        full_text = ' '.join(text_content)
        print(f"\nWord document content length: {len(full_text)} characters")
    except Exception as e:
        print(f"Error processing Word document: {e}")
        return f"Error processing Word document: {e}"

    # Process all codes from Excel
    print("\nProcessing Excel codes...")
    for code in codes_from_excel:
        normalized_code = normalize_code(code)
        print(f"Processing code: {code} (normalized: {normalized_code})")
        
        # First check single_code_sales
        if normalized_code in single_code_sales:
            mapped_code = single_code_sales[normalized_code]
            results.append((code, mapped_code))
            print(f"Added from single_code_sales: {code} -> {mapped_code}")
        # Then check group_code_sales
        elif normalized_code in group_code_sales:
            mapped_code = group_code_sales[normalized_code]
            results.append((code, mapped_code))
            print(f"Added from group_code_sales: {code} -> {mapped_code}")
        else:
            # If no mapping found, keep original code
            results.append((code, code))
            print(f"Added original code: {code} -> {code}")

    # Load the existing Excel workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    # Update Sales Code column (Column E)
    sales_code_col = 5  # Column E is the 5th column (1-based index)
    for row in range(2, ws.max_row + 1):  # Assuming row 1 is header
        wers_code_cell = ws.cell(row=row, column=3)  # Assuming WERS Code is in Column C (3rd column)
        sales_code_cell = ws.cell(row=row, column=sales_code_col)
        
        if wers_code_cell.value:
            wers_code = normalize_code(str(wers_code_cell.value))
            # First check single_code_sales
            if wers_code in single_code_sales:
                sales_code_cell.value = single_code_sales[wers_code]
            # Then check group_code_sales
            elif wers_code in group_code_sales:
                sales_code_cell.value = group_code_sales[wers_code]

    # Save the updated Excel file
    updated_excel_filename = 'updated_' + excel_file.filename
    updated_excel_path = os.path.join(uploads_dir, updated_excel_filename)
    wb.save(updated_excel_path)

    # Debugging: Print final results
    print(f"\nFinal Results (Total: {len(results)}):")
    for wers_code, sales_code in results:
        print(f"{wers_code} -> {sales_code}")

    return render_template('results.html', results=results, updated_excel_path=updated_excel_filename)

@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join('uploads', filename)
    print(f"Attempting to download: {file_path}")  # Debugging
    if not os.path.exists(file_path):
        return f"File not found: {file_path}", 404
    return send_file(file_path, as_attachment=True)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
